import datetime
import time
import openpyxl
import pytest
from fpdf import FPDF
from selenium import webdriver
import allure
from selenium.common.exceptions import TimeoutException
from selenium.webdriver.common.by import By
from selenium.webdriver.support.wait import WebDriverWait
from selenium.webdriver.support import expected_conditions as EC

@allure.step("Entering username ")
def enter_username(username):
  driver.find_element_by_id("un").send_keys(username)

@allure.step("Entering password ")
def enter_password(password):
  driver.find_element_by_id("pw").send_keys(password)

@pytest.fixture()
def test_setup():
  global driver
  global TestName
  global description
  global TestResult
  global TestResultStatus
  global TestDirectoryName
  global checkcount
  global path
  global Exe
  global Dict
  global Dict2
  global FundsNamesList

  TestName = "test_DeleteAtDetailedCAPSTMTs"
  description = "Test case to verify Delete error in Detailed Cap Stmts inside Beneficient"
  TestResult = []
  TestResultStatus = []
  TestFailStatus = []
  FundsNamesList = []
  Dict = {}
  Dict2 = {}

  FailStatus = "Pass"
  TestDirectoryName = "test_Regression"
  Exe = "Yes"
  Directory = 'test_Regression/'
  path = 'C:/BIDS/beneficienttest/Beneficient/' + Directory

  ExcelFileName = "Execution"
  locx = (path + 'Executiondir/' + ExcelFileName + '.xlsx')
  wbx = openpyxl.load_workbook(locx)
  sheetx = wbx.active
  for ix in range(1, 100):
    if sheetx.cell(ix, 1).value == None:
      break
    else:
      if sheetx.cell(ix, 1).value == TestName:
        if sheetx.cell(ix, 2).value == "No":
          Exe = "No"
        elif sheetx.cell(ix, 2).value == "Yes":
          Exe = "Yes"

  if Exe == "Yes":
    driver=webdriver.Chrome(executable_path="C:/BIDS/beneficienttest/Beneficient/Chrome/chromedriver.exe")
    driver.implicitly_wait(10)
    driver.maximize_window()
    driver.get("https://beneficienttest.appiancloud.com/suite/")
    enter_username("neeraj.kumar")
    enter_password("Crochet@786")
    driver.find_element_by_xpath("//input[@type='submit']").click()

  yield
  if Exe == "Yes":
    ct = datetime.datetime.now().strftime("%d_%B_%Y_%I_%M%p")
    ctReportHeader = datetime.datetime.now().strftime("%d %B %Y %I %M%p")

    class PDF(FPDF):
      def header(self):
        self.image(path + 'EmailReportContent/Ben.png', 10, 8, 33)
        self.set_font('Arial', 'B', 15)
        self.cell(73)
        self.set_text_color(0, 0, 0)
        self.cell(35, 10, ' Test Report ', 1, 1, 'B')
        self.set_font('Arial', 'I', 10)
        self.cell(150)
        self.cell(30, 10, ctReportHeader, 0, 0, 'C')
        self.ln(20)

      def footer(self):
        self.set_y(-15)
        self.set_font('Arial', 'I', 8)
        self.set_text_color(0, 0, 0)
        self.cell(0, 10, 'Page ' + str(self.page_no()) + '/{nb}', 0, 0, 'C')

    pdf = PDF()
    pdf.alias_nb_pages()
    pdf.add_page()
    pdf.set_font('Times', '', 12)
    pdf.cell(0, 10, "Test Case Name:  " + TestName, 0, 1)
    pdf.multi_cell(0, 10, "Description:  " + description, 0, 1)

    for i in range(len(TestResult)):
      pdf.set_fill_color(255, 255, 255)
      pdf.set_text_color(0, 0, 0)
      if (TestResultStatus[i] == "Fail"):
        pdf.set_text_color(255, 0, 0)
        TestFailStatus.append("Fail")
      TestName1 = TestResult[i].encode('latin-1', 'ignore').decode('latin-1')
      pdf.multi_cell(0, 7, str(i + 1) + ")  " + TestName1, 0, 1, fill=True)
      TestFailStatus.append("Pass")
    pdf.output(TestName + "_" + ct + ".pdf", 'F')

    # -----------To check if any failed Test case present------------------
    for io in range(len(TestResult)):
      if TestFailStatus[io] == "Fail":
        FailStatus = "Fail"
    # ---------------------------------------------------------------------

    # -----------To add test case details in PDF details sheet-------------
    ExcelFileName = "FileName"
    loc = (path + 'PDFFileNameData/' + ExcelFileName + '.xlsx')
    wb = openpyxl.load_workbook(loc)
    sheet = wb.active
    print()
    check = TestName
    PdfName = TestName + "_" + ct + ".pdf"
    checkcount = 0

    for i in range(1, 100):
      if sheet.cell(i, 1).value == None:
        if checkcount == 0:
          sheet.cell(row=i, column=1).value = check
          sheet.cell(row=i, column=2).value = PdfName
          sheet.cell(row=i, column=3).value = TestDirectoryName
          sheet.cell(row=i, column=4).value = description
          sheet.cell(row=i, column=5).value = FailStatus
          checkcount = 1
        wb.save(loc)
        break
      else:
        if sheet.cell(i, 1).value == check:
          if checkcount == 0:
            sheet.cell(row=i, column=2).value = PdfName
            sheet.cell(row=i, column=3).value = TestDirectoryName
            sheet.cell(row=i, column=4).value = description
            sheet.cell(row=i, column=5).value = FailStatus
            checkcount = 1
    # ----------------------------------------------------------------------------

    # ---------------------To add Test name in Execution sheet--------------------
    ExcelFileName1 = "Execution"
    loc1 = (path + 'Executiondir/' + ExcelFileName1 + '.xlsx')
    wb1 = openpyxl.load_workbook(loc1)
    sheet1 = wb1.active
    checkcount1 = 0

    for ii1 in range(1, 100):
      if sheet1.cell(ii1, 1).value == None:
        if checkcount1 == 0:
          sheet1.cell(row=ii1, column=1).value = check
          checkcount1 = 1
        wb1.save(loc1)
        break
      else:
        if sheet1.cell(ii1, 1).value == check:
          if checkcount1 == 0:
            sheet1.cell(row=ii1, column=1).value = check
            checkcount1 = 1
    # -----------------------------------------------------------------------------

    driver.quit()

@pytest.mark.regression
def test_NavToDetailedCAPSTMTs(test_setup):
  if Exe == "Yes":
      try:
          SHORT_TIMEOUT = 5
          LONG_TIMEOUT = 400
          LOADING_ELEMENT_XPATH = "//div[@id='appian-working-indicator-hidden']"

          # ---------------------------Verify Liquid Trusts page-----------------------------
          PageName = "Funds"
          PageTitle = "Funds - BIDS"
          driver.find_element_by_xpath("//*[@title='" + PageName + "']").click()
          start = time.time()
          try:
              WebDriverWait(driver, SHORT_TIMEOUT
                            ).until(EC.presence_of_element_located((By.XPATH, LOADING_ELEMENT_XPATH)))

              WebDriverWait(driver, LONG_TIMEOUT
                            ).until_not(EC.presence_of_element_located((By.XPATH, LOADING_ELEMENT_XPATH)))
          except TimeoutException:
              pass
          try:
              time.sleep(2)
              bool1 = driver.find_element_by_xpath(
                "//div[@class='appian-context-ux-responsive']/div[4]/div/div/div[1]").is_displayed()
              if bool1 == True:
                ErrorFound1 = driver.find_element_by_xpath(
                  "//div[@class='appian-context-ux-responsive']/div[4]/div/div/div[1]").text
                print(ErrorFound1)
                driver.find_element_by_xpath(
                  "//div[@class='appian-context-ux-responsive']/div[4]/div/div/div[2]/div/button").click()
                TestResult.append(PageName + " not able to open\n" + ErrorFound1)
                TestResultStatus.append("Fail")
                bool1 = False
                driver.close()
          except Exception:
              try:
                time.sleep(2)
                bool2 = driver.find_element_by_xpath(
                  "//div[@class='MessageLayout---message MessageLayout---error']").is_displayed()
                if bool2 == True:
                  ErrorFound2 = driver.find_element_by_xpath(
                    "//div[@class='MessageLayout---message MessageLayout---error']/div/p").text
                  print(ErrorFound2)
                  TestResult.append(PageName + " not able to open\n" + ErrorFound2)
                  TestResultStatus.append("Fail")
                  bool2 = False
                  driver.close()
              except Exception:
                  pass
              pass
          time.sleep(1)
          try:
              assert PageTitle in driver.title, PageName + " not able to open"
              TestResult.append(PageName + " page Opened successfully")
              TestResultStatus.append("Pass")
          except Exception:
              TestResult.append(PageName + " page not able to open")
              TestResultStatus.append("Fail")
          stop = time.time()
          TimeString = stop - start
          print("The time of the run for " + PageName + " is: ", stop - start)
          print(TimeString)

          # --------------------Clicking on A Fund--------------
          PageName = "A Fund"
          Ptitle1 = "User Input Task - BIDS"
          driver.find_element_by_xpath("//tbody/tr[1][@class='PagingGridLayout---selectable']/td[2]").click()
          start = time.time()
          try:
            WebDriverWait(driver, SHORT_TIMEOUT
                          ).until(EC.presence_of_element_located((By.XPATH, LOADING_ELEMENT_XPATH)))

            WebDriverWait(driver, LONG_TIMEOUT
                          ).until_not(EC.presence_of_element_located((By.XPATH, LOADING_ELEMENT_XPATH)))
          except TimeoutException:
            pass
          try:
              time.sleep(2)
              bool1 = driver.find_element_by_xpath(
                "//div[@class='appian-context-ux-responsive']/div[4]/div/div/div[1]").is_displayed()
              if bool1 == True:
                  ErrorFound1 = driver.find_element_by_xpath(
                    "//div[@class='appian-context-ux-responsive']/div[4]/div/div/div[1]").text
                  print(ErrorFound1)
                  driver.find_element_by_xpath(
                    "//div[@class='appian-context-ux-responsive']/div[4]/div/div/div[2]/div/button").click()
                  TestResult.append(PageName + " not able to open\n" + ErrorFound1)
                  TestResultStatus.append("Fail")
                  bool1 = False
                  driver.close()
          except Exception:
              try:
                  time.sleep(2)
                  bool2 = driver.find_element_by_xpath(
                    "//div[@class='MessageLayout---message MessageLayout---error']").is_displayed()
                  if bool2 == True:
                    ErrorFound2 = driver.find_element_by_xpath(
                      "//div[@class='MessageLayout---message MessageLayout---error']/div/p").text
                    print(ErrorFound2)
                    TestResult.append(PageName + " not able to open\n" + ErrorFound2)
                    TestResultStatus.append("Fail")
                    bool2 = False
                    driver.close()
              except Exception:
                pass
              pass
          time.sleep(1)
          try:
              PageTitle1=driver.title
              print(PageTitle1)
              assert Ptitle1 in PageTitle1, PageName + " is not able to open successfully"
              TestResult.append(PageName + " opened successfully")
              TestResultStatus.append("Pass")
          except Exception as e1:
              print(e1)
              TestResult.append(PageName + " is not able to open successfully")
              TestResultStatus.append("Fail")
          stop = time.time()
          TimeString = stop - start
          print("The time of the run for " + PageName + " is: ", stop - start)
          print(TimeString)
          print()

          try:
              driver.find_element_by_xpath("//button[text()='Detailed Cap Stmts']").click()
              print("Detailed Cap Stmts tab is clicked")
          except Exception:
              print("Detailed Cap Stmts tab is not clickable")
              allure.attach(driver.get_screenshot_as_png(), name="Image1", attachment_type=allure.attachment_type.PNG)
              pytest.fail("Failed to click on Detailed Cap Stmts tab")

          # --------------------Clicking Detailed Cap Stmts tab--------------
          PageName = "Detailed Cap Stmts tab"
          Ptitle1 = "User Input Task - BIDS"
          driver.find_element_by_xpath("//button[text()='Detailed Cap Stmts']").click()
          start = time.time()
          try:
            WebDriverWait(driver, SHORT_TIMEOUT
                          ).until(EC.presence_of_element_located((By.XPATH, LOADING_ELEMENT_XPATH)))

            WebDriverWait(driver, LONG_TIMEOUT
                          ).until_not(EC.presence_of_element_located((By.XPATH, LOADING_ELEMENT_XPATH)))
          except TimeoutException:
            pass
          try:
            time.sleep(2)
            bool1 = driver.find_element_by_xpath(
              "//div[@class='appian-context-ux-responsive']/div[4]/div/div/div[1]").is_displayed()
            if bool1 == True:
              ErrorFound1 = driver.find_element_by_xpath(
                "//div[@class='appian-context-ux-responsive']/div[4]/div/div/div[1]").text
              print(ErrorFound1)
              driver.find_element_by_xpath(
                "//div[@class='appian-context-ux-responsive']/div[4]/div/div/div[2]/div/button").click()
              TestResult.append(PageName + " not able to open\n" + ErrorFound1)
              TestResultStatus.append("Fail")
              bool1 = False
              driver.close()
          except Exception:
            try:
              time.sleep(2)
              bool2 = driver.find_element_by_xpath(
                "//div[@class='MessageLayout---message MessageLayout---error']").is_displayed()
              if bool2 == True:
                ErrorFound2 = driver.find_element_by_xpath(
                  "//div[@class='MessageLayout---message MessageLayout---error']/div/p").text
                print(ErrorFound2)
                TestResult.append(PageName + " not able to open\n" + ErrorFound2)
                TestResultStatus.append("Fail")
                bool2 = False
                driver.close()
            except Exception:
              pass
            pass
          time.sleep(1)
          try:
            PageTitle1=driver.title
            print(PageTitle1)
            assert Ptitle1 in PageTitle1, PageName + " is not able to open successfully"
            TestResult.append(PageName + " opened successfully")
            TestResultStatus.append("Pass")
          except Exception as e1:
            print(e1)
            TestResult.append(PageName + " is not able to open successfully")
            TestResultStatus.append("Fail")
          stop = time.time()
          TimeString = stop - start
          print("The time of the run for " + PageName + " is: ", stop - start)
          print(TimeString)
          print()

          Check=0
          try:
             if driver.find_element_by_xpath("//a[contains(text(),'Add/Edit Detailed Cap Statement')]").is_displayed():
              Check=1
          except Exception:
             if driver.find_element_by_xpath("//button[text()='Cancel']").is_displayed():
              Check=2

          if   Check==1:
            driver.find_element_by_xpath("//a[contains(text(),'Add/Edit Detailed Cap Statement')]").click()

          elif Check==2:
             driver.find_element_by_xpath("//button[text()='Cancel']").click()
             TestResult.append("Cancel button clicked successfully")
             TestResultStatus.append("Pass")

             driver.find_element_by_xpath("//a[contains(text(),'Add/Edit Detailed Cap Statement')]").click()
             TestResult.append("Add/Edit Detailed Cap Statement clicked successfully")
             TestResultStatus.append("Pass")

          time.sleep(3)
          driver.find_element_by_xpath("//div[@class='ContentLayout---content_layout']/div[4]/div/div/div/div[1]/div[2]/div[2]/div/table/tbody/tr[position()=1]/td[position()=2]/div/input").clear()

          driver.find_element_by_xpath("//div[@class='ContentLayout---content_layout']/div[4]/div/div/div/div[1]/div[2]/div[2]/div/table/tbody/tr[position()=1]/td[position()=2]/div/input").send_keys("12")
          TestResult.append("Data able to add successfully")
          TestResultStatus.append("Pass")
          time.sleep(3)

          # Data = driver.find_element_by_xpath("//td[2][@class='EditableGridLayout---reducedPadding']/div/input[1]").get_attribute("")
          # time.sleep(5)

          driver.find_element_by_xpath("//button[text()='Save']").click()
          TestResult.append("Save button clicked successfully")
          TestResultStatus.append("Pass")
          time.sleep(3)

          driver.find_element_by_xpath("//a[contains(text(),'Add/Edit Detailed Cap Statement')]").click()

          time.sleep(3)
          driver.find_element_by_xpath("//div[@class='ContentLayout---content_layout']/div[4]/div/div/div/div[1]/div[2]/div[2]/div/table/tbody/tr[position()=1]/td[last()]/div/p/a").click()
          TestResult.append("Delete button clicked successfully")
          TestResultStatus.append("Pass")

          driver.find_element_by_xpath("//button[text()='Save']").click()
          time.sleep(5)
          Data=driver.find_element_by_xpath("//div[@class='ContentLayout---content_layout']/div[4]/div/div/div/div[1]/div[2]/div[1]/div/table/tbody/tr[position()=1]/td[position()=2]/p").text

          print("Data is "+Data)
          if Data in "_":
            TestResult.append("Delete functionality is working correctly")
            TestResultStatus.append("Pass")

      except Exception as Mainerror:
          stop = time.time()
          RoundFloatString = round(float(stop - start), 2)
          print("The time of the run for " + PageName + " is: ", RoundFloatString)
          stringMainerror = repr(Mainerror)
          if stringMainerror in "InvalidSessionIdException('invalid session id', None, None)":
             pass
          else:
              TestResult.append(stringMainerror)
              TestResultStatus.append("Fail")
  else:
      print()
      print("Test Case skipped as per the Execution sheet")
      skip = "Yes"

      # -----------To add Skipped test case details in PDF details sheet-------------
      ExcelFileName = "FileName"
      loc = (path + 'PDFFileNameData/' + ExcelFileName + '.xlsx')
      wb = openpyxl.load_workbook(loc)
      sheet = wb.active
      check = TestName

      for i in range(1, 100):
          if sheet.cell(i, 1).value == check:
            sheet.cell(row=i, column=5).value = "Skipped"
            wb.save(loc)
      # ----------------------------------------------------------------------------
